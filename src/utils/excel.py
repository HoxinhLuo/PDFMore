from openpyxl import Workbook, load_workbook


def writeExcel(path, value, sheet):
    """
    :param sheet:sheet的名称
    :param path:文件的名字和路径
    :param value: 写入的数据
    :return:
    """
    book = Workbook()
    sheet1 = book.active
    sheet1.title = sheet

    for i in range(0, len(value)):
        for j in range(0, len(value[i])):
            sheet1.cell(row=i + 1, column=j + 1, value=str(value[i][j]))

    book.save(path)
    print("写入数据成功！")


def addExcel(path, value, sheet):
    """
    :param path: 写入excel的路径
    :param value: 追加的数据
    :param sheet: sheet的名称
    """
    wb = load_workbook(path)
    wb.create_sheet(sheet)
    ws = wb[sheet]

    for ss in value:
        ws.append(ss)
    wb.save(path)
    print("追加写入成功")